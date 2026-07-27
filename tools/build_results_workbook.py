"""
Combine per-class grading results into ONE Excel workbook, one sheet per
class: exports/Exam_B3_U6-10_Results.xlsx (falls back to _v2 if locked).

Reads .tmp/bubble_results_<class>.json for every class listed below.

Run: py tools/build_results_workbook.py
"""

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "exports", "Exam_B3_U6-10_Results.xlsx")

SHEETS = [("PM 91", "pm91"), ("PM 82", "pm82")]

HEAD = ["Student", "Score /40", "%", "Part I /10", "Part II /10",
        "Part III /10", "Part IV /10", "Blank", "Missed questions"]


def main():
    wb = Workbook()
    wb.remove(wb.active)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="8B1F12")

    for label, cls in SHEETS:
        path = os.path.join(ROOT, ".tmp", f"bubble_results_{cls}.json")
        if not os.path.exists(path):
            print(f"skip {label}: no results file")
            continue
        with open(path, encoding="utf-8") as f:
            results = json.load(f)["results"]

        ws = wb.create_sheet(label)
        ws.append(HEAD)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")
        for r in results:
            ws.append([r["name"], r["total"], r["pct"], r["part_I"], r["part_II"],
                       r["part_III"], r["part_IV"], r["blank"],
                       " ".join(map(str, r["missed"]))])
        n = len(results)
        avg = sum(r["total"] for r in results) / n
        ws.append([])
        ws.append([f"Class average: {avg:.1f} / 40 ({avg/40*100:.0f}%)  ·  {n} students"])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["I"].width = 40
        for col in "BCDEFGH":
            ws.column_dimensions[col].width = 11
        ws.freeze_panes = "A2"

    out = OUT
    try:
        wb.save(out)
    except PermissionError:
        out = OUT.replace(".xlsx", "_v2.xlsx")
        wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
